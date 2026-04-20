# app/core/audit_export.py
"""Чтение JSON-аудита по дням и генерация отчётов Excel / PDF / CSV."""

from __future__ import annotations

import csv as csv_module
import io
import json
import logging
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, AsyncIterator, Iterable

import aiofiles
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.config import Settings

logger = logging.getLogger(__name__)

DEJAVU_FONT_CANDIDATES: tuple[Path, ...] = (
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf"),
    Path("/usr/share/fonts/TTF/DejaVuSans.ttf"),
)


@dataclass
class AuditExportStats:
    total: int = 0
    success: int = 0
    failed: int = 0

    def add_entry(self, entry: dict[str, Any]) -> None:
        self.total += 1
        if entry.get("success") is True:
            self.success += 1
        else:
            self.failed += 1


def resolve_pdf_font_path(settings: Settings) -> Path:
    raw = getattr(settings, "audit_export_pdf_font_path", None)
    if raw:
        p = Path(raw)
        if p.is_file():
            return p
        logger.warning("AUDIT_EXPORT_PDF_FONT_PATH не найден: %s", p)
    for candidate in DEJAVU_FONT_CANDIDATES:
        if candidate.is_file():
            return candidate
    raise FileNotFoundError(
        "Не найден шрифт DejaVuSans для PDF. Установите fonts-dejavu или задайте "
        "AUDIT_EXPORT_PDF_FONT_PATH."
    )


def format_status(success: Any) -> str:
    if success is True:
        return "Успех"
    if success is False:
        return "Неудача"
    return str(success)


def format_extra_column(entry: dict[str, Any]) -> str:
    md = entry.get("metadata")
    bucket: dict[str, Any] = {}
    if isinstance(md, dict):
        bucket.update(md)
    fn = entry.get("filename")
    if fn:
        bucket.setdefault("filename", fn)
    reason = entry.get("reason")
    if reason:
        bucket.setdefault("reason", reason)
    try:
        return json.dumps(bucket, ensure_ascii=False)
    except (TypeError, ValueError):
        return str(bucket)


def shorten_cell(text: str, max_len: int = 2000) -> str:
    t = text.replace("\r", " ").replace("\n", " ")
    if len(t) <= max_len:
        return t
    return t[: max_len - 1] + "…"


# Единый заголовок для всех форматов экспорта.
AUDIT_ROW_HEADERS: list[str] = [
    "Дата", "Пользователь", "Действие", "IP", "Статус", "Доп.данные",
]


def audit_row(entry: dict[str, Any]) -> list[str]:
    """Сформировать строку аудита в едином порядке колонок для всех форматов."""
    return [
        entry.get("timestamp", ""),
        entry.get("user", ""),
        entry.get("action", ""),
        entry.get("ip", ""),
        format_status(entry.get("success")),
        format_extra_column(entry),
    ]


async def iter_audit_log_entries(
    logs_dir: Path,
    start_date: date,
    end_date: date,
    user_id: str | None,
    event_type: str | None,
) -> AsyncIterator[dict[str, Any]]:
    """Построчное чтение audit_YYYY-MM-DD.log за интервал дат (фильтры по user / action)."""
    root = Path(logs_dir)
    if not root.is_dir():
        return

    day = start_date
    one = timedelta(days=1)
    while day <= end_date:
        log_path = root / f"audit_{day:%Y-%m-%d}.log"
        day += one
        if not log_path.is_file():
            continue

        try:
            async with aiofiles.open(log_path, encoding="utf-8") as af:
                while True:
                    line = await af.readline()
                    if not line:
                        break
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry: dict[str, Any] = json.loads(line)
                    except json.JSONDecodeError:
                        logger.debug("Пропуск некорректной JSON-строки в %s", log_path)
                        continue

                    if user_id is not None and str(entry.get("user", "")) != str(user_id):
                        continue
                    if event_type is not None and entry.get("action") != event_type:
                        continue

                    yield entry
        except OSError as e:
            logger.error("Ошибка чтения %s: %s", log_path, e)


async def load_filtered_audit_entries(
    settings: Settings,
    start_date: date,
    end_date: date,
    user_id: str | None,
    event_type: str | None,
) -> tuple[list[dict[str, Any]], AuditExportStats]:
    rows: list[dict[str, Any]] = []
    stats = AuditExportStats()
    logs_dir = Path(settings.audit_logs_dir)
    async for entry in iter_audit_log_entries(
        logs_dir, start_date, end_date, user_id, event_type
    ):
        rows.append(entry)
        stats.add_entry(entry)
    return rows, stats


def build_excel_bytes(
    rows: Iterable[dict[str, Any]],
    stats: AuditExportStats,
    period_start: date,
    period_end: date,
    user_id: str | None,
    event_type: str | None,
) -> bytes:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_sum.append(["Параметр", "Значение"])
    ws_sum.append(["Дата начала периода", period_start.isoformat()])
    ws_sum.append(["Дата окончания периода", period_end.isoformat()])
    ws_sum.append(["Фильтр: пользователь", user_id or "—"])
    ws_sum.append(["Фильтр: тип события (действие)", event_type or "—"])
    ws_sum.append([])
    ws_sum.append(["Всего событий", stats.total])
    ws_sum.append(["Успешных", stats.success])
    ws_sum.append(["Неудачных", stats.failed])

    ws_det = wb.create_sheet("Детали", 1)
    ws_det.append(AUDIT_ROW_HEADERS)
    for entry in rows:
        ws_det.append(audit_row(entry))

    ws_det.auto_filter.ref = ws_det.dimensions
    ws_det.freeze_panes = "A2"

    for col_idx in range(1, ws_det.max_column + 1):
        max_length = 10
        letter = get_column_letter(col_idx)
        for row in ws_det.iter_rows(min_row=1, max_row=ws_det.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
        ws_det.column_dimensions[letter].width = min(max_length + 2, 80)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(
    settings: Settings,
    rows: list[dict[str, Any]],
    stats: AuditExportStats,
    period_start: date,
    period_end: date,
    user_id: str | None,
    event_type: str | None,
) -> bytes:
    font_path = resolve_pdf_font_path(settings)
    if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans", str(font_path)))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="Аудит SMDG",
    )

    title_style = ParagraphStyle(
        "Title",
        fontName="DejaVuSans",
        fontSize=16,
        leading=20,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "Body",
        fontName="DejaVuSans",
        fontSize=9,
        leading=11,
    )
    small_style = ParagraphStyle(
        "Small",
        fontName="DejaVuSans",
        fontSize=8,
        leading=10,
    )

    story: list[Any] = []
    story.append(Paragraph("Аудит SMDG", title_style))
    story.append(
        Paragraph(
            f"<b>Период:</b> с {period_start.isoformat()} по {period_end.isoformat()}",
            body_style,
        )
    )
    filt_parts = [
        f"<b>Пользователь:</b> {user_id or '—'}",
        f"<b>Тип события:</b> {event_type or '—'}",
    ]
    story.append(Paragraph(" &nbsp;|&nbsp; ".join(filt_parts), small_style))
    story.append(Spacer(1, 10))

    # Соответствует порядку AUDIT_ROW_HEADERS: Дата, Пользователь, Действие, IP, Статус, Доп.данные.
    column_max_lens = [40, 24, 48, 18, None, 120]
    data: list[list[str]] = [list(AUDIT_ROW_HEADERS)]
    for entry in rows:
        row = audit_row(entry)
        data.append([
            shorten_cell(str(cell), max_len) if max_len is not None else str(cell)
            for cell, max_len in zip(row, column_max_lens)
        ])

    tbl = Table(data, repeatRows=1, colWidths=[None, 22 * mm, 42 * mm, 22 * mm, 18 * mm, 52 * mm])
    tbl.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), "DejaVuSans", 7),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8E8E8")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 14))
    story.append(
        Paragraph(
            f"<b>Итого:</b> событий — {stats.total}; успешных — {stats.success}; "
            f"неудачных — {stats.failed}",
            body_style,
        )
    )

    doc.build(story)
    return buf.getvalue()


def iter_csv_chunks(rows: Iterable[dict[str, Any]]) -> Iterable[bytes]:
    """Генератор байтовых фрагментов CSV (;, utf-8-sig) для StreamingResponse."""
    sio = io.StringIO()
    writer = csv_module.writer(sio, delimiter=";")
    writer.writerow(AUDIT_ROW_HEADERS)
    yield sio.getvalue().encode("utf-8-sig")

    for entry in rows:
        sio.seek(0)
        sio.truncate(0)
        writer.writerow(audit_row(entry))
        yield sio.getvalue().encode("utf-8")
